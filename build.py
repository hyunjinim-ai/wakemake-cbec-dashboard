#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
웨이크메이크 중국 역직구 대시보드 — 데이터 빌드 스크립트 (v2)
================================================================
데이터 소스(구글시트 / 원본 엑셀) → data.json → index.html 자동 생성.
3개 뷰: ① 일자별 매출(브랜드별)  ② 플랫폼 통합(티몰·도우인·샤오홍슈)  ③ 티몰 월별(기존).

사용법:
  python build.py --gsheet [시트URL|ID]      # ★구글시트('중국역직구 플랫폼 판매데이터') → 티몰 월별+일자별 동기화(매월)
  python build.py --seed                    # 확보된 원본에서 실데이터로 data.json 시드(최초 1회)
  python build.py --sheet [sheet.config.json]   # 구글시트(내 스키마 탭별 공개 CSV) → data.json 재생성
  python build.py --render-only             # data.json 만 바꾼 뒤 index.html 재생성
  python build.py --seed-csv                # 현재 data.json → 구글시트 붙여넣기용 스타터 CSV 출력
  python build.py <티몰원본.xlsx>            # (기존) 티몰 6시트 엑셀 → 티몰 파트 재계산
  python build.py --from-raw <index_2.html>  # (기존) RAW 대시보드에서 티몰 파트 복원

환율 1元 = 220원 · 매출은 증정품 제외 결제금액 기준.
"""
import sys, os, json, re, csv, io, datetime

try:
    sys.stdout.reconfigure(encoding="utf-8")  # Windows 콘솔(cp949) 한글/기호 출력 보장
except Exception:
    pass

from mapping import MAPPING, AMBIGUOUS, norm_cn

FX = 220
HERE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(HERE, "template.html")
DATA_JSON = os.path.join(HERE, "data.json")
OUT_HTML = os.path.join(HERE, "index.html")
SHEET_CFG = os.path.join(HERE, "sheet.config.json")
SEED_DIR = os.path.join(HERE, "sheets_seed")
# 도우인 일자별 + 샤오홍슈 聚光 원본(로컬, 시드 전용 · 커밋 안 함)
RAW_MULTI = r"C:\Users\user\Downloads\로우 확인중.xlsx"

# ---------- 상품 한글명 매핑 (없으면 코드 뒤 6자리) ----------
NAME_MAP = {
 "802291666650":"소프트 블러링 아이팔레트","916056804484":"워터 글로우 코팅밤(구형)",
 "979107752691":"워터 글로우 코팅밤AD(쿠션)","802325822279":"믹스 블러링 볼륨 쉐딩",
 "937100575838":"[임박] 워터 글로우 코팅 쿠션","991053980918":"디파이닝 커버 컨실러",
 "944149219488":"워터풀 글로우 틴트","923641001774":"와이드 파운데이션 브러시",
 "981293620333":"쉐이킹 블러 치크","834612380188":"[한정특가] 워터 글로우 코팅 쿠션",
 "943692099726":"[임박] 디파이닝 커버 컨실러","930412381224":"워터 글로우 코팅 쿠션",
 "1025567503245":"[기획] 디파이닝 커버 컨실러","956917552339":"헬시 글로우 밤 스틱",
 "979918030260":"소프트 블러링 밤 스틱","944132243773":"워터 블러링 레이어 틴트",
 "1056920657682":"볼드 립 블러 틴트","1055872138792":"스테이 픽서 세팅 미스트",
 "930390492188":"소프트 블러링 아이팔레트",
 "955175851181":"디파이닝 커버 파운데이션(브러시증정)","1010846509316":"워터 글로우 코팅 쿠션(브러시증정)",
 "970326378569":"워터풀 글로우 틴트(신규)","994394931225":"WAKEMAKE唯可魅眉笔染眉膏 野生眉",
 "915599957043":"6색 멀티 팔레트","955630790429":"심리스 웨어 쿠션",
 "923655776577":"【特价】WAKEMAKE唯可魅露水美光唇釉","802226967628":"윤곽 형광 조색 팔레트",
 "802231575085":"정형 컨실러 조색 팔레트","802234879735":"퓨어 소이 파우더 팩트",
}
# 유입경로 세부 채널 한글명 (l4 코드 → 한글)
TR = {"搜索":"검색","推荐":"추천 피드","淘宝直播":"타오바오 라이브","天猫榜单":"티몰 랭킹",
      "淘金币及相关场域":"타오진비(포인트)","一淘":"이타오(캐시백)","关键词推广":"키워드 광고(무계)",
      "淘宝客":"타오바오커(제휴)","购物车":"장바구니","我的淘宝":"마이 타오바오",
      "手淘拍立淘":"이미지 검색","淘口令分享":"외부 공유 링크","消息":"메시지 알림",
      "天猫APP":"티몰 APP","逛逛":"타오바오 SNS 피드","自主访问":"직접 방문"}
GMAP = {"经营优势":"자연 유입(무료)","付费推广":"유료 광고","主动回访":"자발적 재방문"}

# 마케팅비 항목 → 플랫폼 귀속(플랫폼 통합 뷰 비용 분해용)
COST_PLATFORM = {
    "티몰 유료광고(내부광고 소모액)": "티몰",
    "도우인 라이브방송(Still 부담 수수료)": "도우인",
    "라이브방송 올리브영 지원분": "도우인",
    "샤오홍씽 CID 효과광고(聚光)": "샤오홍슈",
    "KOL/KOC/SNS 바이럴 마케팅": "샤오홍슈",
}

def monthNum(m):
    mm = re.sub(r"[^0-9]", "", str(m))
    return int(mm) if mm else 0
_LINEMAP = {}   # 티몰글로벌 시트 라인명(id→국문명, 브랜드접두 제거)
def set_linemap(d):
    global _LINEMAP; _LINEMAP = d or {}
_NAME_FIX = {   # 시트 라인명 오기 교정(시트 수정 시 시트 값이 우선하므로 자동 무해화)
    "탕후루 탱글 꿀로스": "탕후루 딥글레이즈 틴트",
}
def kname(pid, cn):
    pid = str(pid)
    nm = _LINEMAP.get(pid)
    if nm and re.search(r"[가-힣]", nm): name = nm          # 시트 국문명 우선(정본)
    elif NAME_MAP.get(pid): name = NAME_MAP[pid]
    else:
        cn = str(cn or "").strip()
        name = cn[:26] if cn else ("코드 " + pid[-6:])      # 국문 매칭 없으면 중문명
    return _NAME_FIX.get(name, name)
def _is_loopy(name):
    n = str(name).lower()
    return "loopy" in n or "루피" in n

def _merge_sellout(products, total_payCNY):
    """동일 제품(같은 라인명) SKU 합산 · loopy 콜라보는 별도 유지."""
    from collections import OrderedDict
    g = OrderedDict(); out = []
    for p in products:
        if _is_loopy(p["name"]): out.append(p); continue
        k = p["name"]
        if k not in g:
            e = dict(p); e["ids"] = [p["id"]]; e["_uvconv"] = (p["uv"] or 0) * (p["conv"] or 0); g[k] = e
        else:
            e = g[k]
            e["payKRW"] += p["payKRW"]; e["payCNY"] += p["payCNY"]; e["uv"] += p["uv"]
            e["_uvconv"] += (p["uv"] or 0) * (p["conv"] or 0); e["ids"].append(p["id"])
            if p["status"] == "온라인": e["status"] = "온라인"
            if p.get("new"): e["new"] = True
    for e in g.values():
        e["conv"] = round(e["_uvconv"] / e["uv"], 2) if e["uv"] else 0
        e.pop("_uvconv", None); out.append(e)
    for p in out:
        p["share"] = round(p["payCNY"] / total_payCNY * 1000) / 10 if total_payCNY else 0
    out.sort(key=lambda x: -x["payKRW"])
    return out

def _merge_ad(adlist):
    """동일 제품(같은 상품명) 광고비 합산 · loopy 콜라보는 별도 유지."""
    from collections import OrderedDict
    g = OrderedDict(); out = []
    for a in adlist:
        if _is_loopy(a["name"]): out.append(a); continue
        k = a["name"]
        if k not in g:
            e = dict(a); e["ids"] = [a["id"]]; g[k] = e
        else:
            e = g[k]
            for f in ("adKRW", "directGmvKRW", "imp", "clk"): e[f] = (e.get(f) or 0) + (a.get(f) or 0)
            e["ids"].append(a["id"])
    for e in g.values():
        e["roi"] = round(e["directGmvKRW"] / e["adKRW"], 2) if e.get("adKRW") else 0
        out.append(e)
    out.sort(key=lambda x: -x["adKRW"])
    return out

def is_gift(p):
    cn = str(p.get("cn") or "")
    if ("赠品" in cn) and (p.get("payCNY") or 0) == 0: return True
    if cn.startswith("【赠품】") or cn.startswith("【赠品】"): return True
    return False
def is_new(p):
    return "新品" in str(p.get("cn") or "")
def numf(v):
    try: return float(v)
    except (TypeError, ValueError): return 0.0

# =========================================================================
#  A. 티몰 원본 엑셀 파서 (원본 대시보드 parseWorkbook 의 Python 이식)  — 기존 유지
# =========================================================================
def parse_workbook(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    def sheet(name):
        if name not in wb.sheetnames: return []
        return [list(r) for r in wb[name].iter_rows(values_only=True)]
    return _parse_tmall_sheets(sheet)

def _parse_tmall_sheets(sheet):
    """티몰 시트들(엑셀 or 구글시트 CSV) → 티몰 파트 원자료. sheet(name) -> rows(list of list)."""
    def norm(m):
        return re.sub(r"\s", "", "" if m is None else str(m)).replace("月", "월")
    def num(v):
        if v is None: return 0.0
        if isinstance(v, (int, float)): return float(v)
        s = str(v).replace(",", "").replace("%", "").replace("元", "").replace("₩", "").strip()
        try: return float(s)
        except ValueError: return 0.0
    def pct(v):  # 전환율: 분수(0.05)·퍼센트("5%"/5) 혼용 흡수 → 퍼센트 숫자
        s = str(v).strip()
        if s.endswith("%"): return round(num(s), 2)
        f = num(s)
        return round(f * 100, 2) if 0 < f <= 1.5 else round(f, 2)

    tmall_rows = []   # '티몰글로벌' 탭: 첫 '티몰' 행=웨메(원), 둘째=컬러그램(백만원)
    for row in sheet("티몰글로벌"):
        if len(row) > 8 and str(row[8]).strip() == "티몰":
            tmall_rows.append([num(x) for x in row[9:15]])
    tmall_actual = tmall_rows[0] if tmall_rows else []
    targets = {"웨이크메이크": tmall_actual}
    if len(tmall_rows) >= 2:   # 컬러그램 섹션은 백만원 → 원 환산
        targets["컬러그램"] = [v * 1e6 if (v and abs(v) < 1e5) else v for v in tmall_rows[1]]
    linemap = {}   # 티몰코드→라인명(국문 정본): 인접셀 스캔 · 한글(국문)이면 항상 우선
    for row in sheet("티몰글로벌"):
        for j in range(len(row) - 1):
            pid = re.sub(r"[^0-9]", "", str(row[j]))
            nmv = str(row[j + 1]).strip()
            if len(pid) >= 11 and nmv and re.search(r"[^\d,.\s]", nmv):
                nmv = re.sub(r"^(웨이크메이크|컬러그램|WAKEMAKE|COLORGRAM)\s*", "", nmv, flags=re.I).strip()
                if not nmv: continue
                if re.search(r"[가-힣]", nmv) or pid not in linemap:   # 국문 우선(중문 선점 방지)
                    linemap[pid] = nmv
    set_linemap(linemap)

    def idx(header, name):
        try: return header.index(name)
        except ValueError:
            for i, h in enumerate(header):
                if str(h).strip() == name: return i
            return -1
    def cell(r, header, name):
        i = idx(header, name)
        return r[i] if (0 <= i < len(r)) else None

    ps = sheet("웨이크메이크_상품별판매데이터")
    products = {}
    if ps:
        h = ps[0]
        for r in ps[1:]:
            if not r: continue
            m = norm(cell(r, h, "통계 일자"))
            if not m or m == "null": continue
            pid = cell(r, h, "상품 ID")
            if pid is None: continue
            pid = str(int(pid) if isinstance(pid, float) else pid)
            products.setdefault(m, []).append({
                "id": pid, "cn": cell(r, h, "상품명"), "status": cell(r, h, "상품 상태"),
                "brand": str(cell(r, h, "브랜드") or "").strip().upper(),
                "uv": int(num(cell(r, h, "상품 방문자 수"))),
                "cart": int(num(cell(r, h, "장바구니 담기 인원 수"))),
                "payCNY": round(num(cell(r, h, "결제 완료 금액"))),
                "conv": pct(cell(r, h, "상품 결제 전환율")),
                "sConv": pct(cell(r, h, "검색 유입 결제 전환율")),
                "sUV": int(num(cell(r, h, "검색 유입 방문자 수"))),
            })

    ts = sheet("웨이크메이크_티몰 점포 유입 현황")
    traffic = {}
    if ts:
        h = ts[0]
        for r in ts[1:]:
            if not r: continue
            m = norm(cell(r, h, "일자"))
            if not m or m == "null": continue
            vis = int(num(cell(r, h, "방문자 수"))); buy = int(num(cell(r, h, "결제 완료 구매자 수")))
            traffic.setdefault(m, []).append({
                "l1": cell(r, h, "1차 유입 경로"), "l2": cell(r, h, "2차 유입 경로"),
                "l3": cell(r, h, "3차 유입 경로"), "l4": cell(r, h, "4차 유입 경로"),
                "visitors": vis, "buyers": buy,
                "conv": round(buy / vis * 10000) / 100 if vis else 0})

    tmall_ad = {}
    ads = sheet("웨이크메이크_티몰 내부 광고 현황")
    if ads:
        h = ads[0]
        for r in ads[1:]:
            if not r: continue
            if "COLORGRAM" in str(cell(r, h, "상품명") or "").upper().replace(" ", ""): continue  # 티몰 유료광고 집계는 WM만(② 호환)
            m = norm(cell(r, h, "통계 일자"))
            if m and m != "null":
                tmall_ad[m] = tmall_ad.get(m, 0) + num(cell(r, h, "광고 소모액(마케팅 비용)"))

    def p2m(p):
        if not p: return None
        s = str(p).split("-")[0].strip()
        mm = re.sub(r"[^0-9].*$", "", s.split(".")[0])
        return (mm + "월") if mm else None
    live_still, live_oy = {}, {}
    ls = sheet("라이브방송 광고 내역")
    if ls:
        h = ls[0]
        for r in ls[1:]:
            if not r: continue
            if str(cell(r, h, "브랜드") or "").lower() != "wakemake": continue
            m = p2m(cell(r, h, "기간 범위"))
            if not m: continue
            live_still[m] = live_still.get(m, 0) + num(cell(r, h, "Still 부담 수수료"))
            live_oy[m] = live_oy.get(m, 0) + num(cell(r, h, "올리브영 지원 내역"))

    xhs_cost, xhs_contrib = {}, {}
    xs = sheet("샤오홍씽-광고")
    if xs:
        h = xs[0]
        for r in xs[1:]:
            if not r: continue
            m = norm(cell(r, h, "월"))
            if not m or m == "null": continue
            c = xhs_cost.setdefault(m, {"ad": 0, "kk": 0})
            c["ad"] += num(cell(r, h, "효과광고(聚光) 비용"))
            c["kk"] += num(cell(r, h, "KOL 마케팅비")) + num(cell(r, h, "KOC 마케팅비")) + num(cell(r, h, "SNS 마케팅비"))
            k = xhs_contrib.setdefault(m, {"visit": 0, "newUV": 0, "buy": 0, "gmv": 0, "ad": 0})
            k["visit"] += num(cell(r, h, "매장 방문 UV")); k["newUV"] += num(cell(r, h, "매장 신규 방문자수"))
            k["buy"] += num(cell(r, h, "전체 매장 구매 UV")); k["gmv"] += num(cell(r, h, "전체 매장 GMV(元)"))
            k["ad"] += num(cell(r, h, "효과광고(聚光) 비용"))

    months = sorted(products.keys(), key=monthNum)
    cost = {}
    for m in months:
        cost[m] = []
        if tmall_ad.get(m): cost[m].append({"owner": "스틸", "item": "티몰 유료광고(내부광고 소모액)", "cny": round(tmall_ad[m])})
        if live_still.get(m): cost[m].append({"owner": "스틸", "item": "도우인 라이브방송(Still 부담 수수료)", "cny": round(live_still[m])})
        if xhs_cost.get(m, {}).get("ad"): cost[m].append({"owner": "OY", "item": "샤오홍씽 CID 효과광고(聚光)", "cny": round(xhs_cost[m]["ad"])})
        if xhs_cost.get(m, {}).get("kk"): cost[m].append({"owner": "OY", "item": "KOL/KOC/SNS 바이럴 마케팅", "cny": round(xhs_cost[m]["kk"])})
        if live_oy.get(m): cost[m].append({"owner": "OY", "item": "라이브방송 올리브영 지원분", "cny": round(live_oy[m])})

    monthly = []
    for i, m in enumerate(months):
        rows = products.get(m, [])
        monthly.append({"month": m, "salesCNY": sum(p["payCNY"] for p in rows),
                        "targetKRW": round(tmall_actual[i]) if i < len(tmall_actual) else 0,
                        "uv": sum(p["uv"] for p in rows), "pv": 0})
    xhs = {}
    for m in months:
        c = xhs_contrib.get(m)
        if not c: continue
        total_uv = next((x["uv"] for x in monthly if x["month"] == m), 0)
        xhs[m] = {"adKRW": round(c["ad"] * FX), "storeVisitUV": round(c["visit"]), "newUV": round(c["newUV"]),
                  "buyUV": round(c["buy"]), "gmvKRW": round(c["gmv"] * FX),
                  "shareOfTotalUV": round(c["visit"] / total_uv * 1000) / 10 if total_uv else 0,
                  "roas": round(c["gmv"] / c["ad"] * 10) / 10 if c["ad"] else 0}
    return {"months": months, "monthly": monthly, "cost": cost, "products": products,
            "traffic": traffic, "xhs": xhs, "targets": targets, "linemap": linemap}

# =========================================================================
#  A-2. 구글시트 판독기 (탭명=엑셀 시트명) + 티몰 일자별  (신규)
# =========================================================================
def _sid(url_or_id):
    m = re.search(r"/spreadsheets/d/([A-Za-z0-9_\-]+)", str(url_or_id))
    return m.group(1) if m else str(url_or_id).strip()

def gsheet_reader(url_or_id, gidmap=None):
    """구글시트(공개) → sheet(name)->rows. gidmap 있으면 export(gid)로(레이트리밋 회피),
    없으면 gviz(탭명)로 읽어 캐시."""
    import urllib.request, urllib.parse
    sid = _sid(url_or_id); cache = {}; gidmap = gidmap or {}
    def sheet(name):
        if name in cache: return cache[name]
        if name in gidmap:
            rows = _fetch_gid_rows(sid, gidmap[name])
        else:
            url = "https://docs.google.com/spreadsheets/d/%s/gviz/tq?tqx=out:csv&sheet=%s" % (sid, urllib.parse.quote(name))
            try:
                req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
                raw = urllib.request.urlopen(req, timeout=30).read().decode("utf-8-sig", "replace")
                rows = list(csv.reader(io.StringIO(raw)))
            except Exception:
                rows = []
        cache[name] = rows
        return rows
    return sheet

def fetch_tmall_ad_products(sid, gid):
    """티몰 내부광고 탭(상품별 소모액·ROI) → {월: [{id,name,brand,adKRW,directGmvKRW,roi,imp,clk}]}."""
    rows = _fetch_gid_rows(sid, gid)
    if not rows: return {}
    h = [str(c).strip() for c in rows[0]]
    def ci(*names):
        for n in names:
            if n in h: return h.index(n)
        return -1
    iM, iId, iNm = ci("통계 일자"), ci("상품 ID"), ci("상품명")
    iAd = ci("광고 소모액(마케팅 비용)", "광고 소모액")
    iGmv, iRoi, iImp, iClk = ci("직접 유입 거래 금액"), ci("광고 직접 ROI"), ci("노출량"), ci("클릭수")
    idx = build_map_index()
    out = {}
    for r in rows[1:]:
        m = re.sub(r"\s", "", str(r[iM]) if 0 <= iM < len(r) else "").replace("月", "월")
        if not m or m in ("None", "null"): continue
        pid = re.sub(r"\.0$", "", str(r[iId]).strip()) if 0 <= iId < len(r) else ""
        if not pid: continue
        ad = _numc(r[iAd]) if 0 <= iAd < len(r) else 0
        if ad <= 0: continue
        nm = str(r[iNm]) if 0 <= iNm < len(r) else ""
        brand = "컬러그램" if "COLORGRAM" in nm.upper().replace(" ", "") else "웨이크메이크"
        out.setdefault(m, []).append({
            "id": pid, "name": kname(pid, nm), "brand": brand,
            "adKRW": round(ad * FX),
            "directGmvKRW": round((_numc(r[iGmv]) if 0 <= iGmv < len(r) else 0) * FX),
            "roi": round(_numc(r[iRoi]) if 0 <= iRoi < len(r) else 0, 2),
            "imp": int(_numc(r[iImp]) if 0 <= iImp < len(r) else 0),
            "clk": int(_numc(r[iClk]) if 0 <= iClk < len(r) else 0),
        })
    for m in out: out[m].sort(key=lambda x: -x["adKRW"])
    return out

def parse_tmall_daily(sheet, sheet_name="웨이크메이크_티몰 일자별매출"):
    """티몰 일자별 매출 탭 → 브랜드별 일자 series (현재 시트엔 WAKEMAKE海外旗舰店 = 웨이크메이크만)."""
    rows = sheet(sheet_name)
    if not rows: return {}
    h = rows[0]
    def ci(*names):
        for n in names:
            if n in h: return h.index(n)
        return -1
    iDate, iStore = ci("통계 일자", "일자"), ci("상점명", "상점", "스토어")
    iPay, iBuy, iOrd = ci("결제 완료 금액"), ci("결제 완료 구매자 수"), ci("결제 완료 하위 주문 건수", "주문 수량")
    iUV = ci("상품 방문자 수(UV)", "방문자 수")
    def num(v):
        s = str(v).replace(",", "").replace("元", "").strip()
        try: return float(s)
        except ValueError: return 0.0
    out = {}
    for r in rows[1:]:
        if iDate < 0 or iDate >= len(r): continue
        d = str(r[iDate]).strip()[:10]
        if not re.match(r"\d{4}-\d{2}-\d{2}", d): continue
        store = str(r[iStore]) if 0 <= iStore < len(r) else ""
        brand = "컬러그램" if ("COLORGRAM" in store.upper() or "colorgram" in store) else "웨이크메이크"
        pay = num(r[iPay]) if 0 <= iPay < len(r) else 0
        out.setdefault(brand, []).append({
            "date": d, "salesKRW": round(pay * FX),
            "orders": int(num(r[iOrd])) if 0 <= iOrd < len(r) else 0,
            "buyers": int(num(r[iBuy])) if 0 <= iBuy < len(r) else 0,
            "uv": int(num(r[iUV])) if 0 <= iUV < len(r) else 0,
        })
    for b in out: out[b].sort(key=lambda x: x["date"])
    return out

# --- 신규 시트 탭: 도우인 월별/상품 · OY마케팅비 (2026-07) ---
_DY_BRAND = {"WAKEMAKE": "웨이크메이크", "COLORGRAM": "컬러그램"}
def _mm_ko(s):
    s = re.sub(r"[^0-9]", "", str(s or ""))
    return (str(int(s[4:6])) + "월") if len(s) >= 6 else ""
def _dy_clean(cn):
    s = re.sub(r"^(WAKE ?MAKE\s*/?\s*唯可魅|韩国\s*colorgram|colorgram|唯可魅)", "", str(cn or ""), flags=re.I)
    s = re.sub(r"【[^】]*】", "", s)
    return s.strip() or str(cn or "")
# 도우인 중문 상품명 → 티몰 한글명 근사 매칭(키워드, 정확도 우선 순서로 배치)
_DY_MATCH_RULES = {
    "웨이크메이크": [
        (["十六色"], "소프트 블러링 아이팔레트"), (["六色"], "6색 멀티 팔레트"),
        (["三色", "修容"], "믹스 블러링 볼륨 쉐딩"), (["三色", "阴影"], "믹스 블러링 볼륨 쉐딩"),
        (["高光修容"], "윤곽 형광 조색 팔레트"),
        (["粉底刷"], "스파츌라 와이드 파운데이션 브러시"),
        (["裸感", "气垫"], "심리스 웨어 쿠션"), (["气垫"], "워터 글로우 코팅 쿠션"),
        (["水光", "唇"], "워터풀 글로우 틴트"), (["玻璃唇"], "워터풀 글로우 틴트"),
        (["叠", "唇"], "워터 블러링 레이어 틴트"),
        (["雾面", "口红"], "볼드 립 블러 틴트"), (["雾面", "唇"], "볼드 립 블러 틴트"),
        (["腮红"], "쉐이킹 블러 치크"), (["眉笔"], "내추럴 하드 브로우 펜슬"),
    ],
    "컬러그램": [
        (["糖葫芦", "精华"], "탕후루 립 세럼"),
        (["糖葫芦", "琥珀"], "탕후루 워터 틴트"), (["糖葫芦", "琉光"], "탕후루 워터 틴트"),
        (["糖葫芦", "水光"], "탕후루 워터 틴트"), (["糖葫芦", "镜面"], "탕후루 워터 틴트"),
        (["糖葫芦"], "탕후루 딥글레이즈 틴트"),
        (["小年糕"], "누디 블러 틴트 (loopy 콜라보)"),
        (["橡皮擦"], "긱 누드 컬러 커버 틴트"), (["唇部打底"], "긱 누드 컬러 커버 틴트"),
        (["卧蚕笔"], "올인원 애교살 메이커"), (["修容笔"], "입체 창조 쉐딩 스틱"),
        (["唇线笔"], "올인원립라이너"), (["眼影盘"], "포켓링 아이 팔레트"),
        (["蜜糖罐"], "쥬시 잼 블러 틴트"), (["果酱"], "틴토리 잼"),
        (["化妆包"], "루피 PVC 파우치"),
    ],
}
def _dy_match(cn, brand):
    s = str(cn or "")
    for kws, ko in _DY_MATCH_RULES.get(brand, []):
        if all(k in s for k in kws): return ko
    return None

def parse_douyin_sheet(sid, gid_daily, gid_prod=None):
    """도우인 일자별매출(载体=全部·시간=不限=일총합)·상품판매(载体=全部) → 브랜드별 월집계+6월 상품."""
    monthly = {}; daily = []; carriers = {}
    _CAR = {"直播": "라이브방송", "商品卡": "상품카드", "短视频": "숏폼영상", "图文": "이미지·글", "其他": "기타"}
    rows = _fetch_gid_rows(sid, gid_daily) if gid_daily else []
    for r in rows[1:] if rows else []:
        if len(r) < 10: continue
        if str(r[4]).strip() != "不限": continue   # 시간대=不限(총합)만
        b = _DY_BRAND.get(str(r[0]).strip().upper()); ko = _mm_ko(r[2])
        if not b or not ko: continue
        carrier = str(r[3]).strip()
        if carrier == "全部":
            d = monthly.setdefault(b, {}).setdefault(ko, {"salesKRW": 0.0, "gmvKRW": 0.0, "orders": 0, "buyers": 0, "impUV": 0, "clickUV": 0})
            d["salesKRW"] += _numc(r[6]) * FX; d["gmvKRW"] += _numc(r[5]) * FX
            d["orders"] += int(_numc(r[7])); d["buyers"] += int(_numc(r[8]))
            d["impUV"] += int(_numc(r[24])) if len(r) > 24 else 0
            d["clickUV"] += int(_numc(r[25])) if len(r) > 25 else 0
            ymd = re.sub(r"[^0-9]", "", str(r[2]))[:8]
            if len(ymd) == 8:
                daily.append({"date": "%s-%s-%s" % (ymd[:4], ymd[4:6], ymd[6:8]),
                              "store": str(r[1]).strip(), "brand": b, "stype": "cross",
                              "gmvCNY": _numc(r[5]), "payCNY": _numc(r[6]),
                              "orders": int(_numc(r[7])), "buyers": int(_numc(r[8])),
                              "aov": _numc(r[9]), "expUV": 0, "clickUV": 0})
        elif ko == "6월":   # 6월 载体(채널/유형)별 점유
            cn = _CAR.get(carrier, carrier)
            cd = carriers.setdefault(b, {})
            cd[cn] = cd.get(cn, 0.0) + _numc(r[6]) * FX
    for b in monthly:
        for ko, d in monthly[b].items():
            d["salesKRW"] = round(d["salesKRW"]); d["gmvKRW"] = round(d["gmvKRW"])
            d["aov"] = round(d["salesKRW"] / d["buyers"]) if d["buyers"] else 0
    products = {}
    if gid_prod:
        prows = _fetch_gid_rows(sid, gid_prod)
        agg = {}
        for r in prows[1:] if prows else []:
            if len(r) < 7 or str(r[5]).strip() != "全部": continue
            b = _DY_BRAND.get(str(r[0]).strip().upper())
            if not b or _mm_ko(r[2]) != "6월": continue
            raw = str(r[3]).strip()
            if "测试" in raw or "勿拍" in raw: continue   # 테스트 상품 제외
            nm = _dy_clean(raw); pay = _numc(r[6]) * FX
            conv = _numc(r[8]) if len(r) > 8 else 0    # 商品점击-成交转化율(구매전환)
            g = agg.setdefault(b, {}).setdefault(nm, {"pay": 0.0, "convw": 0.0})
            g["pay"] += pay; g["convw"] += conv * pay
        for b in agg:
            grp = {}   # 티몰 한글명 매칭 시 동일명끼리 합산(티몰 sell-out 방식과 일관)
            for nm, v in agg[b].items():
                if v["pay"] <= 0: continue
                ko = _dy_match(nm, b); disp = ko or nm
                g = grp.setdefault(disp, {"name": disp, "matched": bool(ko), "payKRW": 0.0, "convw": 0.0, "cns": []})
                g["payKRW"] += v["pay"]; g["convw"] += v["convw"]; g["cns"].append(nm)
            tot = sum(g["payKRW"] for g in grp.values())
            lst = [{"name": g["name"], "matched": g["matched"], "payKRW": round(g["payKRW"]),
                    "share": round(g["payKRW"] / tot * 1000) / 10 if tot else 0,
                    "conv": round(g["convw"] / g["payKRW"] * 10000) / 100 if g["payKRW"] else 0,
                    "cn": " · ".join(g["cns"][:3])} for g in grp.values()]
            lst.sort(key=lambda x: -x["payKRW"])
            products[b] = lst
    daily.sort(key=lambda x: (x["brand"], x["date"]))
    carriers = {b: {k: round(v) for k, v in sorted(c.items(), key=lambda x: -x[1])} for b, c in carriers.items()}
    return {"monthly": monthly, "products": products, "daily": daily, "carriers": carriers}

def parse_oy_mkt(sid, gid):
    """OY마케팅 비용 사용내역 → {브랜드:{월:{속성:₩}}} (실제 집행 마케팅비, 브랜드별)."""
    B = {"CG": "컬러그램", "WM": "웨이크메이크", "COLORGRAM": "컬러그램", "WAKEMAKE": "웨이크메이크"}
    rows = _fetch_gid_rows(sid, gid); out = {}
    if not rows: return out
    h = [str(c).strip() for c in rows[0]]
    def ci(*ns):
        for x in ns:
            for i, c in enumerate(h):
                if x in c: return i
        return -1
    iM, iB, iAttr = ci("月份"), ci("品牌"), ci("属性")
    iKRW = next((i for i, c in enumerate(h) if "₩" in c and "千万" not in c), -1)
    if min(iM, iB, iAttr, iKRW) < 0: return out
    for r in rows[1:]:
        if len(r) <= iKRW: continue
        b = B.get(str(r[iB]).strip().upper())
        mo = re.search(r"(\d+)\s*月", str(r[iM]))
        if not b or not mo: continue
        ko = mo.group(1) + "월"; cat = str(r[iAttr]).strip() or "기타"
        krw = round(_numc(r[iKRW]))
        if krw <= 0: continue
        out.setdefault(b, {}).setdefault(ko, {})
        out[b][ko][cat] = out[b][ko].get(cat, 0) + krw
    return out

def parse_wm_xhs(sid, gid):
    """'웨이크메이크 샤오홍씽-광고'(WM 소유) 월별 → {월:{jg,inflow,gmv,kol}} KRW.
    A=월 · E(4)=쥐광(聚光) 광고비 · U(20)=샤오홍싱 유입매출 · V(21)=캠페인 GMV · X(23)=KOL 마케팅비."""
    rows = _fetch_gid_rows(sid, gid) if gid else []
    out = {}
    for r in rows[1:] if rows else []:
        if len(r) < 24: continue
        mo = re.search(r"(\d+)", str(r[0]))
        if not mo: continue
        d = out.setdefault(mo.group(1) + "월", {"jg": 0.0, "inflow": 0.0, "gmv": 0.0, "kol": 0.0})
        d["jg"] += _numc(r[4]) * FX; d["inflow"] += _numc(r[20]) * FX
        d["gmv"] += _numc(r[21]) * FX; d["kol"] += _numc(r[23]) * FX
    return {ko: {k: round(v) for k, v in d.items()} for ko, d in out.items()}

def parse_cg_xhs(sid, gid_jg, gid_cid):
    """CG 쥐광(gid_jg='컬러그램 샤오홍씽' 消费)+CID(gid_cid='컬러그램 CID-광고' 消费) 월별 → {월:{jg,cid}} KRW."""
    out = {}
    def add(gid, key):
        rows = _fetch_gid_rows(sid, gid) if gid else []
        if not rows: return
        h = [str(c).strip() for c in rows[0]]
        icost = next((i for i, c in enumerate(h) if "消费" in c), -1)
        imon = next((i for i, c in enumerate(h) if ("月份" in c or c == "월")), 0)
        if icost < 0: return
        for r in rows[1:]:
            if len(r) <= max(icost, imon): continue
            mo = re.search(r"(\d+)", str(r[imon]))
            if not mo: continue
            out.setdefault(mo.group(1) + "월", {"jg": 0.0, "cid": 0.0})[key] += _numc(r[icost]) * FX
    add(gid_jg, "jg"); add(gid_cid, "cid")
    return {ko: {"jg": round(v["jg"]), "cid": round(v["cid"])} for ko, v in out.items()}

def parse_tmall_daily_monthly(sid, gid):
    """'티몰 일자별매출' → {브랜드:{월: 결제완료금액KRW}}. A(0)=일자·B(1)=상점명·Q(16)=결제완료금액."""
    rows = _fetch_gid_rows(sid, gid) if gid else []
    out = {}
    for r in rows[1:] if rows else []:
        if len(r) < 17: continue
        store = str(r[1]).upper()
        b = "웨이크메이크" if "WAKEMAKE" in store else ("컬러그램" if "COLORGRAM" in store else None)
        if not b: continue
        mo = re.search(r"(\d{4})[-/.]?(\d{1,2})", str(r[0]))
        if not mo: continue
        out.setdefault(b, {}).setdefault(int(mo.group(2)), 0.0)
        out[b][int(mo.group(2))] += _numc(r[16]) * FX
    return {b: {("%d월" % m): round(v) for m, v in mm.items()} for b, mm in out.items()}

def parse_live_support(sid, gid):
    """라이브방송 광고내역(抖音) → {브랜드:{월:지원비KRW}} = Σ(협업수수료% × 판매GMV)."""
    B = {"colorgram": "컬러그램", "wakemake": "웨이크메이크", "wake make": "웨이크메이크"}
    rows = _fetch_gid_rows(sid, gid); out = {}
    if not rows: return out
    h = [str(c).strip() for c in rows[0]]
    def ci(*ns):
        for x in ns:
            for i, c in enumerate(h):
                if x in c: return i
        return -1
    iM, iB, iP, iFee, iGmv = ci("라이브 방송 월", "방송 월"), ci("브랜드"), ci("플랫폼"), ci("협업 수수료", "수수료"), ci("판매 GMV", "GMV")
    if min(iM, iB, iFee, iGmv) < 0: return out
    for r in rows[1:]:
        if len(r) <= max(iM, iB, iFee, iGmv): continue
        if iP >= 0 and "抖音" not in str(r[iP]): continue
        bl = str(r[iB]).strip().lower()
        b = next((v for k, v in B.items() if k in bl), None)
        mo = re.search(r"(\d+)", str(r[iM]))
        if not b or not mo: continue
        out.setdefault(b, {}).setdefault(mo.group(1) + "월", 0.0)
        out[b][mo.group(1) + "월"] += _numc(r[iFee]) / 100.0 * _numc(r[iGmv]) * FX
    return {b: {ko: round(v) for ko, v in mm.items()} for b, mm in out.items()}

def apply_tmall_marketing(tmall, wmXhs, cgXhs, adp):
    """WM 쥐광(wmXhs)·CG 쥐광+CID(cgXhs)를 티몰 series에 반영 · paidRoas=매출/(내부광고+쥐광+CID) · CPUV/CTR."""
    adp = adp or {}; wmXhs = wmXhs or {}; cgXhs = cgXhs or {}
    impclk = {}
    for m, lst in adp.items():
        for p in lst:
            b = p.get("brand") or "웨이크메이크"
            d = impclk.setdefault((b, m), {"imp": 0, "clk": 0})
            d["imp"] += p.get("imp", 0); d["clk"] += p.get("clk", 0)
    for bko, bdata in tmall.get("tmallBrands", {}).items():
        for s in bdata["series"]:
            m = s["month"]
            if bko == "웨이크메이크":
                w = wmXhs.get(m, {})
                jg = w.get("jg", 0)
                s["xhsAdKRW"] = jg; s["jgKRW"] = jg; s["cidKRW"] = 0
                s["kolKRW"] = w.get("kol", 0); s["xhsInflowKRW"] = w.get("inflow", 0)
                s["paidAdKRW"] = s.get("tmallAdKRW", 0) + jg
            else:
                jc = cgXhs.get(m, {"jg": 0, "cid": 0})
                xh = jc.get("jg", 0) + jc.get("cid", 0)
                s["xhsAdKRW"] = xh; s["jgKRW"] = jc.get("jg", 0); s["cidKRW"] = jc.get("cid", 0)
                s["paidAdKRW"] = s.get("tmallAdKRW", 0) + xh
            s["paidRoas"] = round(s["salesKRW"] / s["paidAdKRW"] * 100) / 100 if s["paidAdKRW"] else 0
            ic = impclk.get((bko, m), {"imp": 0, "clk": 0})
            s["adImp"] = ic["imp"]; s["adClk"] = ic["clk"]
            s["ctr"] = round(ic["clk"] / ic["imp"] * 1000) / 10 if ic["imp"] else 0
            s["cpuv"] = round(s.get("paidAdKRW", 0) / s["uv"]) if s.get("uv") else 0
    if tmall.get("tmallBrands"):
        tmall["series"] = tmall["tmallBrands"]["웨이크메이크"]["series"]
        tmall["byMonth"] = tmall["tmallBrands"]["웨이크메이크"]["byMonth"]

# =========================================================================
#  B. 기존 대시보드(RAW)에서 티몰 파트 시드 — 엑셀 없이 검증용  — 기존 유지
# =========================================================================
def seed_from_raw(html_path):
    txt = open(html_path, "r", encoding="utf-8").read()
    m = re.search(r"const RAW\s*=\s*`(.+?)`", txt, re.S)
    if not m: raise SystemExit("index_2.html 에서 const RAW 를 찾지 못했습니다")
    d = json.loads(m.group(1))
    def unpack(rows, keys): return [dict(zip(keys, r)) for r in rows]
    products = {mo: unpack(rows, d["pkeys"]) for mo, rows in d["products"].items()}
    traffic = {mo: unpack(rows, d["tkeys"]) for mo, rows in d["traffic"].items()}
    return {"months": d["months"], "monthly": d["monthly"], "cost": d["cost"],
            "products": products, "traffic": traffic, "xhs": d.get("xhs", {})}

# =========================================================================
#  C. 티몰 파트 계산 (series + byMonth)  — 기존 유지
# =========================================================================
_BRAND_RAW = {"웨이크메이크": "WAKEMAKE", "컬러그램": "COLORGRAM"}
def _adbrand(name):
    return "컬러그램" if "COLORGRAM" in str(name).upper().replace(" ", "") else "웨이크메이크"

def compute_tmall(u, ad=None):
    """브랜드별 티몰 series+byMonth. ad={월:[상품별광고…]}(fetch_tmall_ad_products) 옵션.
    반환 {months, series(웨메·②호환), byMonth(웨메), tmallBrands:{브랜드:{series,byMonth}}}."""
    months = u["months"]
    monthly_by = {x["month"]: x for x in u["monthly"]}
    adBM = {m: {"웨이크메이크": [], "컬러그램": []} for m in months}
    for m, lst in (ad or {}).items():
        for p in lst:
            b = p.get("brand") or _adbrand(p.get("name"))
            if b not in ("웨이크메이크", "컬러그램"): b = "웨이크메이크"
            adBM.setdefault(m, {"웨이크메이크": [], "컬러그램": []}).setdefault(b, []).append(p)

    def bprods(m, bkey):
        return [p for p in u["products"].get(m, [])
                if not is_gift(p) and str(p.get("brand") or "WAKEMAKE").upper() == bkey]

    def one_brand(bko, bkey):
        isWM = (bko == "웨이크메이크")
        btargets = u.get("targets", {}).get(bko) or []
        series = []
        for mi, m in enumerate(months):
            rows = bprods(m, bkey)
            sales = sum(p["payCNY"] for p in rows) * FX
            uv = sum(p["uv"] for p in rows)
            adlist = sorted(adBM.get(m, {}).get(bko, []), key=lambda a: -a["adKRW"])
            cs = u["cost"].get(m, []) if isWM else []
            if adlist:
                tmAd = round(sum(a["adKRW"] for a in adlist))
            elif isWM:
                tmAd = round(sum(c["cny"] for c in cs if c["item"] == "티몰 유료광고(내부광고 소모액)") * FX)
            else:
                tmAd = 0
            xh = {}   # 샤오홍슈 聚光/CID = 컬그 채널(웨메 미집행) · 티몰 series에는 미포함(#6 수정)
            xhsAd = round(xh.get("adKRW", 0))
            paid = tmAd + xhsAd
            mo = monthly_by.get(m, {}) if isWM else {}
            target = round(btargets[mi]) if (mi < len(btargets) and btargets[mi]) else (round(mo.get("targetKRW", 0)) if isWM else 0)
            if isWM:
                costK = round(sum(c["cny"] for c in cs) * FX)
                oy = round(sum(c["cny"] for c in cs if c["owner"] == "OY") * FX)
                st = round(sum(c["cny"] for c in cs if c["owner"] == "스틸") * FX)
            else:
                costK, oy, st = paid, xhsAd, tmAd
            series.append({
                "month": m, "salesKRW": round(sales), "targetKRW": target,
                "achv": round(sales / target * 1000) / 10 if target else 0,
                "uv": uv, "pv": mo.get("pv", 0) or 0,
                "costKRW": costK, "oyKRW": oy, "stKRW": st,
                "tmallAdKRW": tmAd, "paidAdKRW": paid,
                "roas": round(sales / costK * 100) / 100 if costK else 0,
                "paidRoas": round(sales / paid * 100) / 100 if paid else 0,
                "xhsAdKRW": xhsAd, "xhsGmvKRW": round(xh.get("gmvKRW", 0)),
                "xhsRoas": xh.get("roas", 0), "xhsShare": xh.get("shareOfTotalUV", 0), "xhsUV": xh.get("storeVisitUV", 0),
            })
        by_month = {}
        for i, m in enumerate(months):
            rows = sorted(bprods(m, bkey), key=lambda p: p.get("payCNY") or 0, reverse=True)
            tot = sum(p["payCNY"] for p in rows)
            products = [{
                "name": kname(p["id"], p.get("cn")), "id": str(p.get("id") or ""),
                "status": "온라인" if p.get("status") == "当前在线" else "품절/내림",
                "new": is_new(p), "uv": p.get("uv") or 0,
                "payKRW": round((p.get("payCNY") or 0) * FX), "payCNY": p.get("payCNY") or 0,
                "conv": p.get("conv") or 0, "share": round((p["payCNY"] / tot * 1000)) / 10 if tot else 0,
            } for p in rows]
            products = _merge_sellout(products, tot)   # 동일제품(같은 라인명) SKU 합산 · loopy 별도
            adlist = _merge_ad(sorted(adBM.get(m, {}).get(bko, []), key=lambda a: -a["adKRW"]))
            if isWM:
                cur = u["cost"].get(m, []); prev = u["cost"].get(months[i - 1], []) if i > 0 else []
                def find(rows_, o, it):
                    return next((c["cny"] * FX for c in rows_ if c["owner"] == o and c["item"] == it), 0)
                keys = []
                for c in cur + prev:
                    if COST_PLATFORM.get(c["item"]) == "샤오홍슈": continue  # 聚光/KOL = 컬그 채널 · 웨메 미집행(#6)
                    k = (c["owner"], c["item"])
                    if k not in keys: keys.append(k)
                cost_items = [{"owner": o, "item": it, "cur": round(find(cur, o, it)),
                               "prev": round(find(prev, o, it)), "diff": round(find(cur, o, it) - find(prev, o, it))}
                              for o, it in keys]
                cost_items.sort(key=lambda x: x["cur"], reverse=True)
                tr = u["traffic"].get(m, []); groups = []
                for g in ["经营优势", "付费推广", "主动回访"]:
                    row = next((r for r in tr if r["l1"] == g and r["l2"] == "汇总"), None)
                    if row: groups.append({"name": GMAP[g], "visitors": row["visitors"], "buyers": row["buyers"], "conv": row["conv"]})
                subs_raw = [r for r in tr if r["l4"] in TR and r["l4"] != "汇总"]
                seen = {}
                for r in subs_raw:
                    nm = TR[r["l4"]]
                    if nm not in seen or r["visitors"] > seen[nm]["visitors"]:
                        seen[nm] = {"name": nm, "visitors": r["visitors"], "buyers": r["buyers"], "conv": r["conv"]}
                subs = sorted(seen.values(), key=lambda x: x["visitors"], reverse=True)[:10]
                traffic = {"groups": groups, "subs": subs}
            else:
                cost_items, traffic = [], {"groups": [], "subs": []}
            by_month[m] = {"products": products, "productTotalKRW": round(tot * FX),
                           "adProducts": adlist, "costItems": cost_items, "traffic": traffic}
        return {"months": months, "series": series, "byMonth": by_month}

    tmallBrands = {bko: one_brand(bko, bkey) for bko, bkey in _BRAND_RAW.items()}
    wm = tmallBrands["웨이크메이크"]
    return {"months": months, "series": wm["series"], "byMonth": wm["byMonth"], "tmallBrands": tmallBrands}

# =========================================================================
#  D. 도우인 · 샤오홍슈 원본 파서  (신규)
# =========================================================================
DOUYIN_STORE = {
    "WAKEMAKE海外旗舰店": ("웨이크메이크", "cross"),
    "WAKE MAKE唯可魅旗舰店": ("웨이크메이크", "domestic"),
    "COLORGRAM海外旗舰店": ("컬러그램", "cross"),
    "COLORGRAM旗舰店": ("컬러그램", "domestic"),
}
def parse_douyin(path):
    """도우인 일자별(스토어x브랜드) 원본 → 정규화 rows."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    if "도우인 웨메 매출" not in wb.sheetnames:
        wb.close(); return []
    ws = wb["도우인 웨메 매출"]
    rows = list(ws.iter_rows(values_only=True)); h = rows[0]
    def ci(name): return h.index(name) if name in h else -1
    iDate, iStore = ci("日期"), ci("载体类型")
    iGmv, iPay, iOrd, iBuy, iAov = ci("成交金额"), ci("用户支付金额"), ci("成交订单数"), ci("成交人数"), ci("客单价")
    iExp, iClk = ci("商品曝光人数"), ci("商品点击人数")
    out = []
    for r in rows[1:]:
        if r[iDate] is None: continue
        store = str(r[iStore])
        brand, stype = DOUYIN_STORE.get(store, (store, "?"))
        out.append({
            "date": str(r[iDate])[:10], "store": store, "brand": brand, "stype": stype,
            "gmvCNY": round(numf(r[iGmv]), 2), "payCNY": round(numf(r[iPay]), 2),
            "orders": int(numf(r[iOrd])), "buyers": int(numf(r[iBuy])), "aov": round(numf(r[iAov]), 2),
            "expUV": int(numf(r[iExp])) if iExp >= 0 else 0, "clickUV": int(numf(r[iClk])) if iClk >= 0 else 0,
        })
    wb.close()
    return out

def parse_xhs_juguang(path):
    """샤오홍슈 聚光(효과광고) 일자·카테고리별 집행 원본 → 월/카테고리 집계."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    if "쥐광 내역" not in wb.sheetnames:
        wb.close(); return {"byMonth": {}, "catByMonth": {}}
    ws = wb["쥐광 내역"]
    rows = list(ws.iter_rows(values_only=True)); h = rows[0]
    def ci(name): return h.index(name) if name in h else -1
    iM, iCost, iExp, iClk, iCat = ci("月份"), ci("消费"), ci("曝光/展现"), ci("点击量"), ci("商品品类")
    byMonth, catByMonth = {}, {}
    for r in rows[1:]:
        if iM < 0 or r[iM] is None: continue
        m = str(int(numf(r[iM]))) if numf(r[iM]) else str(r[iM])
        cost, exp, clk = numf(r[iCost]), int(numf(r[iExp])), int(numf(r[iClk]))
        b = byMonth.setdefault(m, {"cost": 0.0, "exp": 0, "clk": 0})
        b["cost"] += cost; b["exp"] += exp; b["clk"] += clk
        cat = str(r[iCat]) if iCat >= 0 and r[iCat] is not None else "(미지정)"
        cm = catByMonth.setdefault(m, {})
        c = cm.setdefault(cat, {"cost": 0.0, "exp": 0, "clk": 0})
        c["cost"] += cost; c["exp"] += exp; c["clk"] += clk
    wb.close()
    return {"byMonth": byMonth, "catByMonth": catByMonth}

# =========================================================================
#  E. 매핑 인덱스 + 미매칭 판정  (신규)
# =========================================================================
def build_map_index():
    """정규화 별칭 → SKU 인덱스, 티몰ID → SKU 인덱스, 애매 별칭 set."""
    by_alias = {}   # (platform, norm) -> sku
    by_tmall = {}   # tmall id -> sku
    for m in MAPPING:
        for pid in m.get("tmallIds", []):
            by_tmall[str(pid)] = m["sku"]
        for plat, key in (("도우인", "douyin"), ("샤오홍슈", "xhs")):
            for a in m.get(key, []):
                by_alias[(plat, norm_cn(a))] = m["sku"]
    ambiguous = {(a["platform"], norm_cn(a["cn"])): a for a in AMBIGUOUS}
    sku_of = {m["sku"]: m for m in MAPPING}
    return {"alias": by_alias, "tmall": by_tmall, "ambiguous": ambiguous, "sku": sku_of}

def match_platform_name(platform, cn, idx):
    """플랫폼 상품명/카테고리 → (sku or None, ambiguous_info or None)."""
    key = (platform, norm_cn(cn))
    if key in idx["alias"]:
        return idx["alias"][key], None
    if key in idx["ambiguous"]:
        return None, idx["ambiguous"][key]
    return None, None

# =========================================================================
#  F. 일자별(페이지①) 계산  (신규)
# =========================================================================
def compute_daily(douyin, month="6월", store_type="cross"):
    prefix = "2026-%02d" % monthNum(month)
    brands = ["웨이크메이크", "컬러그램"]
    series = {b: [] for b in brands}
    for d in douyin:
        if d["stype"] != store_type: continue
        if not d["date"].startswith(prefix): continue
        if d["brand"] not in series: continue
        series[d["brand"]].append({
            "date": d["date"], "salesKRW": round(d["payCNY"] * FX), "gmvKRW": round(d["gmvCNY"] * FX),
            "orders": d["orders"], "buyers": d["buyers"], "aov": round(d["aov"] * FX),
        })
    for b in brands:
        series[b].sort(key=lambda x: x["date"])
    dates = sorted({x["date"] for b in brands for x in series[b]})
    return {
        "month": month, "brands": brands,
        "platformLabel": "도우인(抖音) 역직구 · 海外旗舰店",
        "metricNote": "일 결제금액(用户支付金额) · 증정 제외 · 환율 220원",
        "coverage": (dates[0] + " ~ " + dates[-1]) if dates else "",
        "dates": dates, "series": series,
        "prev": {b: None for b in brands},   # 5월 일자별 미확보 → 시트 입력 시 채움
    }

def _fetch_gid_rows(sid, gid):
    import urllib.request
    if not gid: return []
    url = "https://docs.google.com/spreadsheets/d/%s/export?format=csv&gid=%s" % (sid, gid)
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        raw = urllib.request.urlopen(req, timeout=30).read().decode("utf-8-sig", "replace")
        return list(csv.reader(io.StringIO(raw)))
    except Exception:
        return []

def _numc(v):
    s = str(v).replace(",", "").replace("元", "").replace("₩", "").replace("%", "").strip()
    if s in ("", "-", "#REF!", "#VALUE!", "None"): return 0.0
    try: return float(s)
    except ValueError: return 0.0

_PLAT_KO = {"TMALL": "티몰", "DOUYIN": "도우인", "XHS": "샤오홍슈", "抖音": "도우인", "小红书": "샤오홍슈"}

def parse_daily_status(cfg):
    """실시간 일자별 시트('26년 하반기 …') → 실데이터 dailyStatus + 월별 플랫폼 추이.
    반환 (dailyStatus, monthlyPlatform) · 실패 시 (None, None)."""
    sid = cfg.get("dailySheetId"); gids = cfg.get("dailyGids", {})
    if not sid or not gids: return None, None, None
    S = _fetch_gid_rows(sid, gids.get("summary"))
    if not S: return None, None, None

    # --- 요약: 브랜드×플랫폼 목표/실적(백만원→원) + 월별 플랫폼 ---
    tgt = {"웨이크메이크": {}, "컬러그램": {}}; act = {"웨이크메이크": {}, "컬러그램": {}}
    cur_b = None
    for r in S:
        def c(i): return str(r[i]).strip() if i < len(r) else ""
        lab = c(3)
        if lab in ("웨이크메이크", "컬러그램"): cur_b = lab; continue
        if cur_b and lab in ("TMALL", "DOUYIN", "XHS"):
            p = _PLAT_KO[lab]
            tgt[cur_b][p] = round(_numc(c(4)) * 1e6)
            act[cur_b][p] = round(_numc(c(5)) * 1e6)
    monthly = {"티몰": {}, "도우인": {}}   # 월별 섹션: 브랜드 idx3, 플랫폼 idx4, 값 idx5+ (WM+CG 합산)
    mbrand = {"웨이크메이크": {"티몰": {}, "도우인": {}}, "컬러그램": {"티몰": {}, "도우인": {}}}  # 브랜드별
    _BKO = {"WAKEMAKE": "웨이크메이크", "COLORGRAM": "컬러그램"}
    for i, r in enumerate(S):
        bkey = str(r[3]).strip() if len(r) > 3 else ""
        if bkey in ("COLORGRAM", "WAKEMAKE"):
            bko = _BKO[bkey]; hdr = S[i]
            cols = [(j, int(re.sub(r"[^0-9]", "", str(hdr[j]))))
                    for j in range(5, len(hdr)) if "월" in str(hdr[j]) and re.sub(r"[^0-9]", "", str(hdr[j]))]
            for r2 in S[i + 1:i + 6]:
                p = str(r2[4]).strip() if len(r2) > 4 else ""
                if p in ("TMALL", "抖音"):
                    ko = "티몰" if p == "TMALL" else "도우인"
                    for cj, mi in cols:
                        v = _numc(r2[cj]) if cj < len(r2) else 0
                        if v > 0:
                            monthly[ko][mi] = monthly[ko].get(mi, 0) + round(v)
                            mbrand[bko][ko][mi] = mbrand[bko][ko].get(mi, 0) + round(v)

    # --- 브랜드 일자별(매출 by 플랫폼·UV·주문) ---
    def parse_brand_daily(gid):
        rows = _fetch_gid_rows(sid, gid)
        hrow = next((i for i, r in enumerate(rows) if any(str(x).strip() == "일자" for x in r)), None)
        if hrow is None: return []
        h = [str(x).strip() for x in rows[hrow]]
        def idx(name, occ=0):
            cnt = 0
            for i, x in enumerate(h):
                if x == name:
                    if cnt == occ: return i
                    cnt += 1
            return -1
        iDate, iTm, iDy, iXh = idx("일자"), idx("티몰글로벌 실적"), idx("도우인 실적"), idx("샤홍슈 실적")
        iUV = [idx("访客数 (UV)", 0), idx("访客数 (UV)", 1)]
        iBuy = [idx("支付人数", 0), idx("支付人数", 1)]
        out = []
        for r in rows[hrow + 1:]:
            if iDate < 0 or iDate >= len(r): continue
            d = str(r[iDate]).strip()[:10]
            if not re.match(r"\d{4}-\d{2}-\d{2}", d): continue
            tm = _numc(r[iTm]) if 0 <= iTm < len(r) else 0
            dy = _numc(r[iDy]) if 0 <= iDy < len(r) else 0
            xh = _numc(r[iXh]) if 0 <= iXh < len(r) else 0
            if tm + dy + xh <= 0: continue
            uv = sum(_numc(r[i]) for i in iUV if 0 <= i < len(r))
            od = sum(_numc(r[i]) for i in iBuy if 0 <= i < len(r))
            out.append({"date": d, "티몰": round(tm), "도우인": round(dy), "샤오홍슈": round(xh),
                        "total": round(tm + dy + xh), "uv": int(uv), "orders": int(od)})
        out.sort(key=lambda x: x["date"])
        return out
    dmap = {"웨이크메이크": parse_brand_daily(gids.get("wakemake")),
            "컬러그램": parse_brand_daily(gids.get("colorgram"))}
    asof_all = [d["date"] for b in dmap for d in dmap[b]]
    if not asof_all: return None, None, None
    asOf = max(asof_all); maxDay = int(asOf[8:10]); mon = int(asOf[5:7])

    # --- UV 전월비(역직구 海外旗舰店 스토어, 당월 MTD vs 전월 동기간) ---
    traf = _fetch_gid_rows(sid, gids.get("traffic"))
    def uv_mom(brand):
        if not traf: return {"cur": 0, "prev": 0}
        h = [str(x).strip() for x in traf[0]]
        iV = h.index("访客数") if "访客数" in h else 2
        key = "WAKEMAKE" if brand == "웨이크메이크" else "COLORGRAM"
        cur = prev = 0
        for r in traf[1:]:
            store = str(r[1]) if len(r) > 1 else ""
            if key not in store.upper().replace(" ", "") or "海外" not in store: continue
            d = str(r[0])[:10]
            if not re.match(r"\d{4}-\d{2}-\d{2}", d): continue
            day = int(d[8:10])
            if day > maxDay: continue
            if int(d[5:7]) == mon: cur += _numc(r[iV])
            elif int(d[5:7]) == mon - 1: prev += _numc(r[iV])
        return {"cur": int(cur), "prev": int(prev)}

    brands = ["웨이크메이크", "컬러그램"]
    data = {b: {"target": tgt[b], "actual": act[b], "uv": uv_mom(b), "daily": dmap[b]} for b in brands}
    ds = {"sample": False, "asOf": asOf, "month": "%d월" % mon, "prevMonth": "%d월" % (mon - 1),
          "dayOfMonth": maxDay, "daysInMonth": 31, "brands": brands,
          "platforms": ["티몰", "도우인", "샤오홍슈"], "data": data}
    mos = sorted(set(list(monthly["티몰"].keys()) + list(monthly["도우인"].keys())))
    mp = {"months": ["%d월" % m for m in mos],
          "티몰": [monthly["티몰"].get(m) for m in mos],
          "도우인": [monthly["도우인"].get(m) for m in mos]}
    tbb = {}   # 티몰 브랜드별 월별 매출 + 7월 실적/목표
    for b in ("웨이크메이크", "컬러그램"):
        tm = mbrand[b]["티몰"]; ms2 = sorted(tm.keys())
        tbb[b] = {"months": ["%d월" % m for m in ms2], "sales": [tm[m] for m in ms2],
                  "target7": tgt[b].get("티몰", 0), "actual7": act[b].get("티몰", 0),
                  "asOfMonth": ds["month"]}
    return ds, mp, tbb

def daily_status_sample():
    """페이지① 상단 '7월 현황' 샘플 — 새 실시간 시트(1B_7…) 연결 전 레이아웃용.
    브랜드×플랫폼 목표·MTD실적·UV(전월 동기 대비). 실데이터 연결 시 이 함수를 파서로 교체."""
    asof, dim = 15, 31
    plats = ["티몰", "도우인", "샤오홍슈"]
    WV = [0.9, 1.0, 1.1, 0.8, 0.72, 1.2, 1.35, 1.05, 0.95, 1.05, 1.18, 0.85, 0.78, 1.12, 1.25]  # 15일 가중(주말 등락)
    def series(mtd):
        s = sum(WV); return [round(mtd * w / s) for w in WV]
    seed = {
        "웨이크메이크": {"target": {"티몰": 155000000, "도우인": 52000000, "샤오홍슈": 21000000},
                    "mtd": {"티몰": 71800000, "도우인": 27300000, "샤오홍슈": 8400000},
                    "uv": {"cur": 63200, "prev": 61050}},
        "컬러그램": {"target": {"티몰": 42000000, "도우인": 34000000, "샤오홍슈": 9000000},
                  "mtd": {"티몰": 16400000, "도우인": 17600000, "샤오홍슈": 2600000},
                  "uv": {"cur": 38900, "prev": 35200}},
    }
    aov = {"웨이크메이크": 24000, "컬러그램": 10500}   # 브랜드 객단가(주문건수 산출용)
    data = {}
    for b, v in seed.items():
        per = {p: series(v["mtd"][p]) for p in plats}
        daily = []
        for i in range(asof):
            day = {"date": "2026-07-%02d" % (i + 1)}
            for p in plats: day[p] = per[p][i]
            day["total"] = sum(day[p] for p in plats)
            day["orders"] = round(day["total"] / aov[b])
            daily.append(day)
        data[b] = {"target": v["target"], "actual": v["mtd"], "uv": v["uv"], "daily": daily}
    return {"sample": True, "asOf": "2026-07-%02d" % asof, "month": "7월", "prevMonth": "6월",
            "dayOfMonth": asof, "daysInMonth": dim, "brands": list(seed.keys()), "platforms": plats, "data": data}

def daily_struct(series_by_brand, label, metricNote, month=None):
    """임의 브랜드별 일자 series → 페이지① 데이터 구조(월 필터 옵션)."""
    brands = ["웨이크메이크", "컬러그램"]
    series = {b: list(series_by_brand.get(b, [])) for b in brands}
    if month:
        pre = "2026-%02d" % monthNum(month)
        series = {b: [x for x in series[b] if x["date"].startswith(pre)] for b in brands}
    for b in brands: series[b].sort(key=lambda x: x["date"])
    dates = sorted({x["date"] for b in brands for x in series[b]})
    if month:
        mlabel = month
    else:
        mos = sorted({int(d[5:7]) for d in dates})
        mlabel = "%d–%d월" % (mos[0], mos[-1]) if len(mos) > 1 else ("%d월" % mos[0] if mos else "전체")
    return {"month": mlabel, "brands": brands, "platformLabel": label,
            "metricNote": metricNote, "coverage": (dates[0] + " ~ " + dates[-1]) if dates else "",
            "dates": dates, "series": series, "prev": {b: None for b in brands}}

# =========================================================================
#  G. 플랫폼 통합(페이지②) 계산  (신규 · 브랜드 차원)
# =========================================================================
WM_BRAND = "웨이크메이크"

def _brand_products(prodmap, brand, platform, idx):
    """옵션 상품별 입력(시트 도우인_상품별/샤오홍슈_상품별) → 매핑 반영 상품 리스트."""
    out = []
    for p in (prodmap or {}).get(brand, []):
        cn = p.get("name") or p.get("cn") or ""
        sku, amb = match_platform_name(platform, cn, idx)
        tid = idx["tmall"].get(str(p.get("tmallId", "")))
        sku = sku or tid
        out.append({"name": cn, "sku": sku, "ko": idx["sku"][sku]["ko"] if sku else cn,
                    "matched": bool(sku), "payKRW": round(numf(p.get("payKRW") or (numf(p.get("payCNY")) * FX))),
                    "orders": int(numf(p.get("orders"))), "uv": int(numf(p.get("uv")))})
    out.sort(key=lambda x: x["payKRW"], reverse=True)
    return out

def compute_platforms(tmall, douyin, xhs, month="6월", douyin_products=None, xhs_products=None):
    """브랜드별(웨이크메이크·컬러그램) 플랫폼 통합 구조."""
    brands = [WM_BRAND, "컬러그램"]
    byBrand = {b: _platforms_for_brand(tmall, douyin, xhs, month, b,
                                       douyin_products=douyin_products, xhs_products=xhs_products) for b in brands}
    return {"brands": brands, "byBrand": byBrand}

def _platforms_for_brand(tmall, douyin, xhs, month, brand, store_type="cross",
                         douyin_products=None, xhs_products=None):
    idx = build_map_index()
    prefix = "2026-%02d" % monthNum(month)
    is_wm = (brand == WM_BRAND)
    tbb = (tmall.get("tmallBrands") or {}).get(brand)   # 브랜드별 티몰(웨메/컬그 모두 확보)
    src_series = tbb["series"] if tbb else tmall["series"]
    src_bm = tbb["byMonth"] if tbb else tmall["byMonth"]
    S = {s["month"]: s for s in src_series}
    BM = src_bm
    s6 = S.get(month, {}); bm6 = BM.get(month, {})
    has_tmall = bool(bm6.get("products"))

    tmall_cost_items = bm6.get("costItems", [])   # WM만 owner/item 구조; CG는 []
    def plat_cost(plat):
        return sum(c["cur"] for c in tmall_cost_items if COST_PLATFORM.get(c["item"]) == plat)
    tm_cost = plat_cost("티몰") or round(s6.get("tmallAdKRW", 0))   # CG는 series의 티몰 유료광고

    # ---- 티몰 ---- (웨메·컬그 모두 브랜드별 로우데이터)
    if has_tmall:
        tmall_p = {
            "label": "티몰글로벌(天猫国际)", "hasData": True, "hasProductDetail": True,
            "kpi": {"salesKRW": s6.get("salesKRW", 0), "uv": s6.get("uv", 0), "costKRW": tm_cost,
                    "roas": round(s6.get("salesKRW", 0) / tm_cost, 2) if tm_cost else 0,
                    "orders": None, "buyers": None},
            "products": bm6.get("products", []), "productTotalKRW": bm6.get("productTotalKRW", 0),
            "costItems": [c for c in tmall_cost_items if COST_PLATFORM.get(c["item"]) == "티몰"],
            "traffic": bm6.get("traffic", {"groups": [], "subs": []}),
            "note": ("상품별 sell-out·유입경로 전체 확보(로우데이터)" if is_wm
                     else "상품별 sell-out·광고 확보 · 유입경로는 시트 미보유"),
        }
    else:
        tmall_p = {
            "label": "티몰글로벌(天猫国际)", "hasData": False, "hasProductDetail": False,
            "kpi": {"salesKRW": 0, "uv": 0, "costKRW": 0, "roas": 0, "orders": None, "buyers": None},
            "products": [], "productTotalKRW": 0, "costItems": [], "traffic": {"groups": [], "subs": []},
            "note": "티몰 " + brand + " 데이터 미확보",
        }

    # ---- 도우인 ---- (두 브랜드 모두 일자별 확보)
    drows = [d for d in douyin if d["stype"] == store_type and d["brand"] == brand and d["date"].startswith(prefix)]
    dsum = {"payKRW": 0, "gmvKRW": 0, "orders": 0, "buyers": 0, "expUV": 0, "clickUV": 0}
    daily = []
    for d in sorted(drows, key=lambda x: x["date"]):
        dsum["payKRW"] += d["payCNY"] * FX; dsum["gmvKRW"] += d["gmvCNY"] * FX
        dsum["orders"] += d["orders"]; dsum["buyers"] += d["buyers"]
        dsum["expUV"] += d["expUV"]; dsum["clickUV"] += d["clickUV"]
        daily.append({"date": d["date"], "salesKRW": round(d["payCNY"] * FX),
                      "orders": d["orders"], "buyers": d["buyers"]})
    dy_cost = plat_cost("도우인")
    dy_sales = round(dsum["payKRW"])
    dprods = _brand_products(douyin_products, brand, "도우인", idx)
    douyin_p = {
        "label": "도우인(抖音) 역직구", "hasData": bool(drows), "hasProductDetail": bool(dprods),
        "kpi": {"salesKRW": dy_sales, "gmvKRW": round(dsum["gmvKRW"]), "costKRW": dy_cost,
                "roas": round(dy_sales / dy_cost, 2) if dy_cost else None,
                "orders": dsum["orders"], "buyers": dsum["buyers"],
                "aov": round(dy_sales / dsum["orders"]) if dsum["orders"] else 0,
                "uv": dsum["expUV"]},
        "products": dprods, "daily": daily,
        "funnel": {"expUV": dsum["expUV"], "clickUV": dsum["clickUV"], "buyers": dsum["buyers"]},
        "note": "스토어 단위 일자 집계" + ("" if dprods else " · 상품별 sell-out 은 시트 `도우인_상품별` 입력 시 표시"),
    }

    # ---- 샤오홍슈 (聚光 집행 + 기여 GMV) ---- (聚光/CID = 컬그 채널 · 웨메 미집행 · #6 수정: 로컬 WM귀속 제거)
    xm = str(monthNum(month))
    xj = {"cost": 0, "exp": 0, "clk": 0}
    xcat = {}
    xhs_ad_juguang = round(xj["cost"] * FX)
    xhs_cost_all = plat_cost("샤오홍슈")  # 聚光 + KOL/KOC/SNS
    xhs_gmv = s6.get("xhsGmvKRW", 0)
    xprods = _brand_products(xhs_products, brand, "샤오홍슈", idx)
    # 카테고리 매핑
    cats, unmatched = [], []
    for cn, v in sorted(xcat.items(), key=lambda x: -x[1]["cost"]):
        sku, amb = match_platform_name("샤오홍슈", cn, idx)
        adK = round(v["cost"] * FX)
        cats.append({"cn": cn, "adKRW": adK, "exp": v["exp"], "clk": v["clk"],
                     "sku": sku, "ko": idx["sku"][sku]["ko"] if sku else None, "matched": bool(sku)})
        if not sku:
            reason = amb["reason"] if amb else "매핑 테이블에 없음 — 상품/콘텐츠 구분 검토"
            cand = [idx["sku"][c]["ko"] for c in (amb["candidates"] if amb else []) if c in idx["sku"]]
            unmatched.append({"platform": "샤오홍슈", "cnName": cn, "metricLabel": "聚光 집행",
                              "valueKRW": adK, "reason": reason, "candidates": cand})
    xhs_p = {
        "label": "샤오홍슈(小红书) 聚光", "hasData": is_wm, "hasProductDetail": bool(xprods),
        "kpi": {"salesKRW": xhs_gmv, "gmvKRW": xhs_gmv, "adKRW": xhs_ad_juguang, "costKRW": xhs_cost_all,
                "roas": round(xhs_gmv / xhs_cost_all, 2) if xhs_cost_all else 0,
                "uv": s6.get("xhsUV", 0), "exp": xj["exp"], "clk": xj["clk"]},
        "categories": cats, "products": xprods,
        "note": ("聚光 로우 기준 집행 상세(카테고리별) · GMV/UV 는 기여 GMV 소스 기준(탭③과 소폭 상이)"
                 if is_wm else "샤오홍슈 " + brand + " 데이터 미확보 — 시트 `플랫폼_샤오홍슈` 입력 대기"),
    }

    byPlatform = {"티몰": tmall_p, "도우인": douyin_p, "샤오홍슈": xhs_p}

    # ---- 통합(합산) ----
    comb_sales = tmall_p["kpi"]["salesKRW"] + douyin_p["kpi"]["salesKRW"] + xhs_p["kpi"]["salesKRW"]
    comb_cost = tm_cost + dy_cost + xhs_cost_all
    by_sales = [{"platform": p, "salesKRW": byPlatform[p]["kpi"]["salesKRW"]} for p in ["티몰", "도우인", "샤오홍슈"]]
    # 매핑 합산 상품(티몰 매출 + 도우인 상품 매출 + 샤오홍슈 확정매칭 聚光비)
    prod_by_sku = {}
    def bucket(sku, ko):
        return prod_by_sku.setdefault(sku, {"sku": sku, "ko": ko, "티몰": 0, "도우인": 0, "샤오홍슈_ad": 0})
    for p in tmall_p["products"]:
        sku = idx["tmall"].get(p.get("id", ""))
        ko = idx["sku"][sku]["ko"] if sku else p["name"]
        bucket(sku or ("TM:" + (p.get("id") or p["name"])), ko)["티몰"] += p["payKRW"]
    for p in douyin_p["products"]:
        sku = p["sku"] or ("DY:" + p["name"])
        bucket(sku, p["ko"])["도우인"] += p["payKRW"]
    for c in cats:
        if c["matched"]:
            bucket(c["sku"], idx["sku"][c["sku"]]["ko"])["샤오홍슈_ad"] += c["adKRW"]
    comb_products = sorted(prod_by_sku.values(), key=lambda x: x["티몰"] + x["도우인"], reverse=True)

    combined = {
        "kpi": {"salesKRW": round(comb_sales), "costKRW": round(comb_cost),
                "roas": round(comb_sales / comb_cost, 2) if comb_cost else 0,
                "orders": douyin_p["kpi"]["orders"]},
        "bySales": by_sales, "products": comb_products,
        "note": "매출 = 티몰 결제(확정) + 도우인 결제(확정) + 샤오홍슈 기여GMV(광고 기여 추정치) · 비용 = 채널 귀속 마케팅비 합. 확정매출과 추정치가 혼합되므로 통합 ROAS는 참고 지표입니다.",
    }

    return {"brand": brand, "month": month, "storeType": store_type,
            "order": ["티몰", "도우인", "샤오홍슈"], "byPlatform": byPlatform,
            "combined": combined, "unmatched": unmatched}

def mapping_out():
    return {"skus": [{"sku": m["sku"], "ko": m["ko"], "brand": m.get("brand", ""),
                      "status": m.get("status", "확정"), "tmallIds": m.get("tmallIds", []),
                      "douyin": m.get("douyin", []), "xhs": m.get("xhs", []),
                      "note": m.get("note", "")} for m in MAPPING],
            "ambiguous": AMBIGUOUS}

# =========================================================================
#  H. 조립 · 렌더 · 시트 어댑터 · 시드 CSV  (신규/확장)
# =========================================================================
def assemble(tmall_u, douyin, xhs, month=None):
    tmall = compute_tmall(tmall_u)
    if month is None:
        month = tmall["months"][-1] if tmall["months"] else "6월"
    data = {"fx": FX}
    data.update(tmall)  # months, series, byMonth
    data["daily"] = compute_daily(douyin, month=month)
    data["platforms"] = compute_platforms(tmall, douyin, xhs, month=month)
    data["mapping"] = mapping_out()
    return data

def render(data):
    if not os.path.exists(TEMPLATE):
        raise SystemExit("template.html 이 없습니다.")
    tpl = open(TEMPLATE, "r", encoding="utf-8").read()
    if "__DATA__" not in tpl:
        raise SystemExit("template.html 에 __DATA__ 자리표시자가 없습니다.")
    html = tpl.replace("__DATA__", json.dumps(data, ensure_ascii=False, separators=(",", ":")))
    open(OUT_HTML, "w", encoding="utf-8").write(html)

def tmall_from_datajson():
    """기존 data.json 에서 티몰 파트 로드 — 원본 엑셀 없이 재조립(브랜드별 포함)."""
    d = json.load(open(DATA_JSON, "r", encoding="utf-8"))
    out = {"months": d["months"], "series": d["series"], "byMonth": d["byMonth"]}
    if d.get("tmallBrands"): out["tmallBrands"] = d["tmallBrands"]
    return out

# ---- 구글시트 어댑터 (탭별 공개 CSV) ----
def fetch_csv(url):
    import urllib.request
    with urllib.request.urlopen(url, timeout=30) as r:
        raw = r.read().decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(raw)))

def _read_products_tab(url):
    """플랫폼_도우인/샤오홍슈 '상품별' 탭 → {brand: [ {name,payCNY,orders,uv,tmallId} ]}."""
    out = {}
    for r in fetch_csv(url):
        name = r.get("상품명(중문)") or r.get("상품명") or r.get("商品名") or r.get("상품") or ""
        if not name: continue
        out.setdefault(r.get("브랜드") or "", []).append({
            "name": name, "tmallId": r.get("티몰_상품ID") or "",
            "payCNY": numf(r.get("결제금액(元)") or r.get("결제금액")), "payKRW": numf(r.get("결제금액(₩)")),
            "orders": numf(r.get("주문수")), "uv": numf(r.get("UV") or r.get("방문자"))})
    return out

def build_from_sheet(cfg_path):
    """sheet.config.json 의 탭별 CSV URL → data.json 재생성.
    신규 데이터(일자별/도우인/샤오홍슈/마케팅/매핑/상품별)를 시트에서, 티몰 파트는
    티몰 탭 미구현 시 기존 data.json 재사용. (스키마는 README 참조)"""
    if not os.path.exists(cfg_path):
        raise SystemExit("설정 파일이 없습니다: %s  (README의 sheet.config.json 예시 참조)" % cfg_path)
    cfg = json.load(open(cfg_path, "r", encoding="utf-8"))
    tabs = cfg.get("tabs", {})
    tmall = tmall_from_datajson()   # 티몰 파트: 기존 유지(시트 티몰 탭 미구현 시)
    # 도우인 일자별
    douyin = []
    if tabs.get("플랫폼_도우인"):
        for r in fetch_csv(tabs["플랫폼_도우인"]):
            store = r.get("스토어", "")
            brand = r.get("브랜드") or DOUYIN_STORE.get(store, ("", ""))[0]
            stype = "cross" if ("海外" in store or r.get("구분") == "역직구") else "domestic"
            douyin.append({"date": str(r.get("일자", ""))[:10], "store": store, "brand": brand, "stype": stype,
                           "gmvCNY": numf(r.get("GMV(元)") or r.get("GMV")), "payCNY": numf(r.get("결제금액(元)") or r.get("결제금액")),
                           "orders": int(numf(r.get("주문수"))), "buyers": int(numf(r.get("구매자수"))),
                           "aov": numf(r.get("객단가")), "expUV": int(numf(r.get("노출UV"))), "clickUV": int(numf(r.get("클릭UV")))})
    # 샤오홍슈 聚光
    xhs = {"byMonth": {}, "catByMonth": {}}
    if tabs.get("플랫폼_샤오홍슈"):
        for r in fetch_csv(tabs["플랫폼_샤오홍슈"]):
            m = re.sub(r"[^0-9]", "", str(r.get("월", "")))
            if not m: continue
            cost, exp, clk = numf(r.get("消费(元)") or r.get("消费")), int(numf(r.get("曝光"))), int(numf(r.get("点击")))
            b = xhs["byMonth"].setdefault(m, {"cost": 0.0, "exp": 0, "clk": 0})
            b["cost"] += cost; b["exp"] += exp; b["clk"] += clk
            cat = r.get("商品品类") or r.get("카테고리") or "(미지정)"
            c = xhs["catByMonth"].setdefault(m, {}).setdefault(cat, {"cost": 0.0, "exp": 0, "clk": 0})
            c["cost"] += cost; c["exp"] += exp; c["clk"] += clk
    dprods = _read_products_tab(tabs["도우인_상품별"]) if tabs.get("도우인_상품별") else None
    xprods = _read_products_tab(tabs["샤오홍슈_상품별"]) if tabs.get("샤오홍슈_상품별") else None
    # 시트에 신규 데이터가 없으면 기존 data.json 의 daily/platforms 유지
    d0 = json.load(open(DATA_JSON, "r", encoding="utf-8")) if os.path.exists(DATA_JSON) else {}
    if not douyin and not xhs["catByMonth"]:
        data = {"fx": FX}; data.update(tmall)
        data["daily"] = d0.get("daily", compute_daily([]))
        data["platforms"] = d0.get("platforms", compute_platforms(tmall, [], xhs, douyin_products=dprods, xhs_products=xprods))
        data["mapping"] = mapping_out()
    else:
        data = assemble_from_parts(tmall, douyin, xhs, douyin_products=dprods, xhs_products=xprods)
    _write_data(data)
    render(data)
    return data

def assemble_from_parts(tmall, douyin, xhs, month=None, douyin_products=None, xhs_products=None):
    if month is None:
        month = tmall["months"][-1] if tmall["months"] else "6월"
    data = {"fx": FX}; data.update(tmall)
    data["daily"] = compute_daily(douyin, month=month)
    data["platforms"] = compute_platforms(tmall, douyin, xhs, month=month,
                                          douyin_products=douyin_products, xhs_products=xhs_products)
    data["mapping"] = mapping_out()
    return data

# ---- 시드 CSV(구글시트 붙여넣기용) ----
def write_seed_csv(data):
    os.makedirs(SEED_DIR, exist_ok=True)
    def w(name, header, rows):
        with open(os.path.join(SEED_DIR, name), "w", encoding="utf-8-sig", newline="") as f:
            wr = csv.writer(f); wr.writerow(header)
            for r in rows: wr.writerow(r)
    # 일자별매출(도우인 역직구, 브랜드별)
    d = data["daily"]; rows = []
    for b in d["brands"]:
        for x in d["series"][b]:
            rows.append([x["date"], b, "도우인", round(x["salesKRW"] / FX, 2), x["orders"], x["buyers"]])
    w("일자별매출.csv", ["일자", "브랜드", "플랫폼", "결제금액(元)", "주문수", "구매자수"], rows)
    # 상품매핑
    rows = [[m["sku"], m["ko"], m.get("brand", ""), ";".join(m.get("tmallIds", [])),
             ";".join(m.get("douyin", [])), ";".join(m.get("xhs", [])), m.get("status", "확정"), m.get("note", "")]
            for m in MAPPING]
    for a in AMBIGUOUS:
        rows.append(["", "", "", "", "", a["cn"], "검토", a["reason"]])
    w("상품매핑.csv", ["SKU코드", "한글상품명", "브랜드", "티몰_상품ID(;)", "도우인_별칭(;)", "샤오홍슈_별칭(;)", "상태", "비고"], rows)
    # 마케팅비(6월, 플랫폼 귀속)
    wm = data["platforms"]["byBrand"][WM_BRAND]; m6 = wm["month"]; bm = data["byMonth"].get(m6, {})
    rows = [[COST_PLATFORM.get(c["item"], "기타"), c["owner"], c["item"], round(c["cur"] / FX), m6]
            for c in bm.get("costItems", [])]
    w("마케팅비.csv", ["플랫폼", "주체", "항목", "금액(元)", "월"], rows)
    # 샤오홍슈 聚光 카테고리
    rows = [[m6, c["cn"], round(c["adKRW"] / FX, 2), c["exp"], c["clk"], "확정" if c["matched"] else "검토"]
            for c in wm["byPlatform"]["샤오홍슈"]["categories"]]
    w("플랫폼_샤오홍슈.csv", ["월", "商品品类", "消费(元)", "曝光", "点击", "매핑상태"], rows)
    # 도우인 일자별(스토어별 원본 형식 · 붙여넣기용)
    dd = data["daily"]
    rows = [[x["date"], b, "海外旗舰店", round(x.get("gmvKRW", x["salesKRW"]) / FX, 2), round(x["salesKRW"] / FX, 2),
             x["orders"], x["buyers"]] for b in dd["brands"] for x in dd["series"][b]]
    w("플랫폼_도우인.csv", ["일자", "브랜드", "스토어", "GMV(元)", "결제금액(元)", "주문수", "구매자수"], rows)
    # 상품별 입력 틀(빈 템플릿 + 예시 1행) — 채우면 상품 단위 통합에 자동 반영
    w("플랫폼_도우인_상품별.csv", ["월", "브랜드", "상품명(중문)", "티몰_상품ID", "결제금액(元)", "주문수", "UV"],
      [[m6, "웨이크메이크", "(예시)水光气垫", "930412381224", "", "", ""]])
    w("플랫폼_샤오홍슈_상품별.csv", ["월", "브랜드", "상품명(중문)", "티몰_상품ID", "결제금액(元)", "주문수", "UV"],
      [[m6, "웨이크메이크", "(예시)粉糖气垫", "930412381224", "", "", ""]])
    print("[시드 CSV] %s 에 생성: 일자별매출 / 플랫폼_도우인 / 플랫폼_샤오홍슈 / 상품매핑 / 마케팅비 / *_상품별(입력틀)" % SEED_DIR)

# =========================================================================
#  I. CLI
# =========================================================================
def _write_data(data):
    """data.json 저장 — 시트 전용 필드(dailyTmall/dailyStatus/monthlyPlatform)는
    다른 모드에서 기존값 보존. dailyStatus 없으면 샘플로 대체."""
    old = {}
    if os.path.exists(DATA_JSON):
        try: old = json.load(open(DATA_JSON, "r", encoding="utf-8"))
        except Exception: old = {}
    for k in ("dailyTmall", "dailyStatus", "monthlyPlatform", "tmallByBrand", "tmallBrands", "douyin", "oyMkt", "wmXhs"):
        if k not in data and old.get(k): data[k] = old[k]
    if "dailyStatus" not in data:
        data["dailyStatus"] = daily_status_sample()
    data["builtAt"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    json.dump(data, open(DATA_JSON, "w", encoding="utf-8"), ensure_ascii=False, indent=1)

def _gsheet_id_from_cfg():
    if os.path.exists(SHEET_CFG):
        try: return json.load(open(SHEET_CFG, "r", encoding="utf-8")).get("gsheetId")
        except Exception: return None
    return None

def _summary(data):
    ms = data["months"]
    print("[완료] 티몰 월: %s (최신 %s)" % (ms, ms[-1] if ms else "-"))
    for key, lab in (("daily", "일자별·도우인"), ("dailyTmall", "일자별·티몰")):
        d = data.get(key)
        if d:
            print("       [%s] %s · %s · %s" % (lab, d["platformLabel"], d["coverage"],
                  " / ".join("%s %d일" % (b, len(d["series"][b])) for b in d["brands"])))
    for brand in data["platforms"]["brands"]:
        pf = data["platforms"]["byBrand"][brand]; um = pf["unmatched"]
        line = "       <%s> " % brand + " · ".join(
            "%s ₩%s" % (p, format(pf["byPlatform"][p]["kpi"].get("salesKRW", 0), ",")) for p in pf["order"])
        print(line)
        print("           통합 매출 ₩%s · ROAS %sx · 미매칭 %d건" % (
              format(pf["combined"]["kpi"]["salesKRW"], ","), pf["combined"]["kpi"]["roas"], len(um)))

def _local_douyin_xhs():
    douyin = parse_douyin(RAW_MULTI) if os.path.exists(RAW_MULTI) else []
    xhs = parse_xhs_juguang(RAW_MULTI) if os.path.exists(RAW_MULTI) else {"byMonth": {}, "catByMonth": {}}
    return douyin, xhs

def main():
    args = sys.argv[1:]
    if not args:
        print(__doc__); sys.exit(0)
    cmd = args[0]

    if cmd == "--render-only":
        data = json.load(open(DATA_JSON, "r", encoding="utf-8"))
        render(data); print("[완료] index.html 재생성"); return

    if cmd == "--seed-csv":
        data = json.load(open(DATA_JSON, "r", encoding="utf-8"))
        write_seed_csv(data); return

    if cmd == "--gsheet":
        # 월별 시트(리뷰·내역) + 실시간 일자별 시트 → 동기화
        cfg = {}
        if os.path.exists(SHEET_CFG):
            try: cfg = json.load(open(SHEET_CFG, "r", encoding="utf-8"))
            except Exception: cfg = {}
        src = args[1] if len(args) > 1 else cfg.get("gsheetId")
        tmall, tmDaily, adp = None, None, None
        if src:
            try:
                sh = gsheet_reader(src, cfg.get("monthlyGids"))
                u = _parse_tmall_sheets(sh)
                if u["months"]:
                    adp = fetch_tmall_ad_products(src, cfg.get("tmallAdGid")) if cfg.get("tmallAdGid") else None
                    tmall = compute_tmall(u, ad=adp)   # 브랜드별 series+byMonth(상품별 광고 포함)
                    tmDaily = parse_tmall_daily(sh)
                    if adp: print("[상품별 광고] %d개월 연결" % len(adp))
            except Exception as e:
                print("[알림] 월별 시트 파싱 오류: %s" % str(e)[:60]); tmall = None
        if tmall is None:
            print("[알림] 월별 시트 읽기 실패/미설정 → 기존 티몰 데이터 유지")
            tmall = tmall_from_datajson()
        # 신규 시트 탭 먼저 파싱 → 도우인 공식 일자별을 ②/① 소스로 승격
        eg = cfg.get("extraGids", {}); dyx, oy, wmXhs, cgXhs, live, tmMon = {}, {}, {}, {}, {}, {}
        if src and eg:
            try:
                _s = _sid(src)
                dyx = parse_douyin_sheet(_s, eg.get("도우인_일자별매출"), eg.get("도우인_상품판매"))
                oy = parse_oy_mkt(_s, eg.get("OY마케팅비용"))
                wmXhs = parse_wm_xhs(_s, eg.get("웨메_샤오홍씽"))                       # WM 쥐광·유입매출·KOL
                cgXhs = parse_cg_xhs(_s, eg.get("컬러그램_샤오홍씽"), eg.get("컬러그램_CID"))  # CG 쥐광+CID
                live = parse_live_support(_s, eg.get("라이브방송"))
                tmMon = parse_tmall_daily_monthly(_s, eg.get("티몰_일자별매출"))          # 매장별 월 실매출
            except Exception as e:
                print("[알림] 신규 탭 파싱 오류: %s" % str(e)[:80]); dyx, oy, wmXhs, cgXhs, live, tmMon = {}, {}, {}, {}, {}, {}
        if tmall and (wmXhs or cgXhs or adp):
            apply_tmall_marketing(tmall, wmXhs, cgXhs, adp)   # WM 쥐광 / CG 쥐광+CID → 티몰 ROAS · CPUV/CTR
            print("[티몰 ROAS] 6월 paidRoas WM=%s / CG=%s · CPUV WM=%s/CG=%s" % (
                next((s.get("paidRoas") for s in tmall["tmallBrands"]["웨이크메이크"]["series"] if s["month"] == "6월"), "-"),
                next((s.get("paidRoas") for s in tmall["tmallBrands"]["컬러그램"]["series"] if s["month"] == "6월"), "-"),
                next((s.get("cpuv") for s in tmall["tmallBrands"]["웨이크메이크"]["series"] if s["month"] == "6월"), "-"),
                next((s.get("cpuv") for s in tmall["tmallBrands"]["컬러그램"]["series"] if s["month"] == "6월"), "-")))
        if wmXhs:
            wm_tm = (tmMon.get("웨이크메이크", {}) or {})
            months = [m for m in (tmall["months"] if tmall else []) if m in wmXhs] or sorted(wmXhs, key=lambda x: int(x[:-1]))
            rows = {}
            for m in months:
                w = wmXhs[m]; jg = w["jg"]; inf = w["inflow"]; kol = w["kol"]; tms = wm_tm.get(m, 0)
                rows[m] = {"jgKRW": jg, "inflowKRW": inf, "gmvKRW": w["gmv"], "kolKRW": kol, "tmallSalesKRW": tms,
                           "inflowShare": round(inf / tms * 1000) / 10 if tms else 0,
                           "jgRoi": round(inf / jg * 100) / 100 if jg else 0,
                           "totalRoi": round(inf / (jg + kol) * 100) / 100 if (jg + kol) else 0}
            data_wmXhs = {"months": months, "rows": rows,
                          "note": "웨이크메이크 샤오홍씽-광고 탭 · 쥐광(聚光)·KOL=OY(올리브영) 집행 · 티몰 실매출=티몰 일자별매출(WAKEMAKE 매장) · 환율 220원"}
        _local_dy, xhs = _local_douyin_xhs()
        douyin = dyx.get("daily") or _local_dy   # 공식 도우인 일자별(전월) 우선, 실패 시 로컬(부분)
        data = assemble_from_parts(tmall, douyin, xhs)
        if dyx.get("monthly"):
            # 도우인 ROI = 매출 / 라이브방송 지원(협업수수료)  [내부광고비 데이터 없음 → 제외]
            for b, mm in dyx["monthly"].items():
                for ko, d in mm.items():
                    ls = live.get(b, {}).get(ko, 0)
                    d["liveKRW"] = ls
                    d["roi"] = round(d["salesKRW"] / ls * 100) / 100 if ls else None
            data["douyin"] = {"monthly": dyx["monthly"], "products": dyx.get("products", {}),
                              "carriers": dyx.get("carriers", {}),
                              "note": "매출=도우인 일자별매출(载体=全部) · ROI=매출/라이브방송지원(협업수수료) · 환율 220원"}
            print("[도우인 시트] 월집계 %s · 6월 상품 %s · 载体 %s · ROI(6월)=%s" %
                  ({b: len(dyx["monthly"][b]) for b in dyx["monthly"]},
                   {b: len(dyx["products"].get(b, [])) for b in dyx["monthly"]},
                   {b: list(dyx.get("carriers", {}).get(b, {}).keys()) for b in dyx["monthly"]},
                   {b: dyx["monthly"][b].get("6월", {}).get("roi") for b in dyx["monthly"]}))
        if oy:
            data["oyMkt"] = oy
            print("[OY마케팅] 브랜드 %s · 6월 %s" %
                  (list(oy.keys()), {b: sum(oy[b].get("6월", {}).values()) for b in oy}))
        if wmXhs:
            data["wmXhs"] = data_wmXhs
            r6 = data_wmXhs["rows"].get("6월", {})
            print("[WM 샤오홍씽] 6월 쥐광 ₩%s · 유입매출 ₩%s(비중 %s%%) · 쥐광ROI %s · 전체ROI %s" %
                  (format(r6.get("jgKRW", 0), ","), format(r6.get("inflowKRW", 0), ","),
                   r6.get("inflowShare"), r6.get("jgRoi"), r6.get("totalRoi")))
        if tmDaily: data["dailyTmall"] = daily_struct(tmDaily, "티몰글로벌(天猫国际) · 海外旗舰店",
                                                     "티몰 일 결제금액 · 증정 제외 · 환율 220원")
        ds, mp, tbb = parse_daily_status(cfg)   # 실시간 일자별 현황(신규 시트)
        if ds:
            data["dailyStatus"] = ds
            if mp and mp["months"]: data["monthlyPlatform"] = mp
            if tbb: data["tmallByBrand"] = tbb
            print("[일자별 현황] 실데이터 연결 · 기준 %s · %s" %
                  (ds["asOf"], " / ".join("%s 실적 ₩%s" % (b, format(sum(ds["data"][b]["actual"].values()), ",")) for b in ds["brands"])))
        else:
            print("[일자별 현황] 새 시트 미설정/읽기 실패 → 기존(또는 샘플) 유지")
        _write_data(data); render(data); _summary(data); return

    if cmd == "--sheet":
        cfg = args[1] if len(args) > 1 else SHEET_CFG
        data = build_from_sheet(cfg); _summary(data); return

    if cmd == "--seed":
        src = args[1] if len(args) > 1 else RAW_MULTI
        if not os.path.exists(src): raise SystemExit("원본이 없습니다: %s" % src)
        tmall = tmall_from_datajson()
        douyin = parse_douyin(src); xhs = parse_xhs_juguang(src)
        data = assemble_from_parts(tmall, douyin, xhs)
        _write_data(data); render(data); _summary(data); return

    if cmd == "--from-raw":
        if len(args) < 2: raise SystemExit("사용법: python build.py --from-raw <index_2.html>")
        tmall = compute_tmall(seed_from_raw(args[1]))
        douyin, xhs = _local_douyin_xhs()
        data = assemble_from_parts(tmall, douyin, xhs)
        _write_data(data); render(data); _summary(data); return

    # 기본: 티몰 원본 엑셀
    path = cmd
    if not os.path.exists(path): raise SystemExit("엑셀 파일을 찾을 수 없습니다: %s" % path)
    tmall = compute_tmall(parse_workbook(path))
    douyin, xhs = _local_douyin_xhs()
    data = assemble_from_parts(tmall, douyin, xhs)
    _write_data(data); render(data); _summary(data)

if __name__ == "__main__":
    main()
